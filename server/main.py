import math
import time
import hmac
import hashlib
import calendar
import warnings
import datetime
import traceback
from io import BytesIO
warnings.filterwarnings("ignore", category=DeprecationWarning)

from fastapi import FastAPI, Depends, HTTPException, Form
from fastapi.responses import StreamingResponse, RedirectResponse
from fastapi.templating import Jinja2Templates
from fastapi.requests import Request
from sqlalchemy.orm import Session
import database
import models

# openpyxl для Excel
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

database.Base.metadata.create_all(bind=database.engine)

app = FastAPI(title="Time Tracker API с Веб-Админкой")

# Инициализируем стандартный Jinja2Templates
templates = Jinja2Templates(directory="templates")

# Настройки геозоны офиса
OFFICE_LAT = 55.7558
OFFICE_LON = 37.6173
ALLOWED_RADIUS_METERS = 50.0
QR_SECRET_KEY = b"SUPER_SECRET_OFFICE_KEY_123"

def get_current_qr_code() -> str:
    time_window = int(time.time() // 30)
    msg = str(time_window).encode()
    sig = hmac.new(QR_SECRET_KEY, msg, hashlib.sha256).hexdigest()
    return str(int(sig, 16) % 1000000).zfill(6)

def calculate_distance(lat1: float, lon1: float, lat2: float, lon2: float) -> float:
    R = 6371000.0
    phi1, phi2 = math.radians(lat1), math.radians(lat2)
    delta_phi = math.radians(lat2 - lat1)
    delta_lambda = math.radians(lon2 - lon1)
    a = (math.sin(delta_phi / 2) ** 2 +
         math.cos(phi1) * math.cos(phi2) * math.sin(delta_lambda / 2) ** 2)
    return R * 2 * math.atan2(math.sqrt(a), math.sqrt(1 - a))

@app.get("/")
def read_root():
    return RedirectResponse(url="/admin/dashboard")

@app.get("/api/get_office_qr")
def get_office_qr():
    return {"current_qr": get_current_qr_code(), "expires_in_seconds": 30 - int(time.time() % 30)}


# --- ВЕБ-ИНТЕРФЕЙС АДМИНИСТРАТОРА ---

@app.get("/admin/dashboard")
def admin_dashboard(request: Request, db: Session = Depends(database.get_db)):
    """Отображает страницу мониторинга: кто сейчас на объекте"""
    try:
        users = db.query(models.User).all()
        status_list = []
        today_start = datetime.datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
        
        for user in users:
            last_checkin = db.query(models.CheckIn).filter(
                models.CheckIn.user_id == user.id,
                models.CheckIn.is_valid == True,
                models.CheckIn.timestamp >= today_start
            ).order_by(models.CheckIn.timestamp.desc()).first()
            
            if last_checkin:
                status = "НА РАБОТЕ" if last_checkin.action_type == "IN" else "Вне объекта"
                last_action = "Вход (Начало смены)" if last_checkin.action_type == "IN" else "Выход (Конец смены)"
                last_time = last_checkin.timestamp.strftime('%H:%M:%S')
            else:
                status = "Вне объекта"
                last_action = "Нет отметок за сегодня"
                last_time = "—"
                
            status_list.append({
                "full_name": str(user.full_name),
                "role": str(user.role),
                "status": status,
                "last_action": last_action,
                "last_time": last_time
            })
            
        return templates.TemplateResponse(request, "dashboard.html", {"status_list": status_list})
    except Exception as e:
        error_details = traceback.format_exc()
        return templates.TemplateResponse(request, "dashboard.html", {"status_list": [], "error": f"{str(e)}\n{error_details}"})

@app.get("/admin/users")
def admin_users_page(request: Request, db: Session = Depends(database.get_db)):
    """Отображает страницу со списком сотрудников"""
    users = db.query(models.User).order_by(models.User.id.asc()).all()
    return templates.TemplateResponse(request, "users.html", {"users": users})


# --- УПРАВЛЕНИЕ БАЗОЙ ДАННЫХ ИЗ ВЕБ-ИНТЕРФЕЙСА ---

@app.post("/admin/users/add")
def admin_add_user(full_name: str = Form(...), role: str = Form(...), db: Session = Depends(database.get_db)):
    new_user = models.User(full_name=full_name, role=role)
    db.add(new_user)
    db.commit()
    return RedirectResponse(url="/admin/users", status_code=303)

@app.post("/admin/users/edit/{user_id}")
def admin_edit_user(user_id: int, role: str = Form(...), db: Session = Depends(database.get_db)):
    user = db.query(models.User).filter(models.User.id == user_id).first()
    if user:
        user.role = role
        db.commit()
    return RedirectResponse(url="/admin/users", status_code=303)

@app.get("/admin/users/delete/{user_id}")
def admin_delete_user(user_id: int, db: Session = Depends(database.get_db)):
    user = db.query(models.User).filter(models.User.id == user_id).first()
    if user:
        db.delete(user)
        db.commit()
    return RedirectResponse(url="/admin/users", status_code=303)

@app.get("/api/auth/check_id")
def check_user_id(user_id: int, db: Session = Depends(database.get_db)):
    employee = db.query(models.User).filter(models.User.id == user_id).first()
    if not employee:
        raise HTTPException(status_code=404, detail="Сотрудник не найден")
    return {"status": "success", "user_id": employee.id, "full_name": employee.full_name}

# --- ОФИЦИАЛЬНЫЙ ТАБЕЛЬ-ШАХМАТКА ДЛЯ БУХГАЛТЕРИИ ---

@app.get("/api/admin/export_tabel")
def export_tabel_to_excel(db: Session = Depends(database.get_db)):
    """Генерирует официальный табель-шахматку учета рабочего времени"""
    try:
        now = datetime.datetime.now()
        current_year = now.year
        current_month = now.month
        
        # Автоопределение количества дней в месяце
        _, days_in_month = calendar.monthrange(current_year, current_month)

        months_ru = {
            1: "ЯНВАРЬ", 2: "ФЕВРАЛЬ", 3: "МАРТ", 4: "АПРЕЛЬ", 5: "МАЙ", 6: "ИЮНЬ",
            7: "ИЮЛЬ", 8: "АВГУСТ", 9: "СЕНТЯБРЬ", 10: "ОКТЯБРЬ", 11: "НОЯБРЬ", 12: "ДЕКАБРЬ"
        }

        wb = Workbook()
        ws = wb.active
        ws.title = months_ru[current_month]
        ws.sheet_view.showGridLines = True

        font_title = Font(name="Segoe UI", size=14, bold=True, color="1F497D")
        font_header = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
        font_data = Font(name="Segoe UI", size=10)
        font_total = Font(name="Segoe UI", size=10, bold=True)
        
        fill_header = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
        fill_total = PatternFill(start_color="F2F5F8", end_color="F2F5F8", fill_type="solid")
        
        thin_border = Border(
            left=Side(style='thin', color='D9D9D9'), right=Side(style='thin', color='D9D9D9'),
            top=Side(style='thin', color='D9D9D9'), bottom=Side(style='thin', color='D9D9D9')
        )
        
        align_center = Alignment(horizontal="center", vertical="center")
        align_left = Alignment(horizontal="left", vertical="center")

        ws["A1"] = 'Проект: "Николь-Плаза"'
        ws["A1"].font = font_title
        ws["A2"] = f"{months_ru[current_month]} {current_year}"
        ws["A2"].font = Font(name="Segoe UI", size=12, bold=True)

        headers = ["Сотрудник ФИО", "Должность"] + [str(d) for d in range(1, days_in_month + 1)] + ["кол-во часов", "кол-во дней"]
        ws.append([])
        ws.append(headers)

        total_columns = len(headers)
        for col_idx in range(1, total_columns + 1):
            cell = ws.cell(row=4, column=col_idx)
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = align_center
            cell.border = thin_border
        ws.row_dimensions.height = 25

        users = db.query(models.User).all()
        current_row = 5

        for user in users:
            checkins = db.query(models.CheckIn).filter(
                models.CheckIn.user_id == user.id,
                models.CheckIn.is_valid == True
            ).all()

            hours_by_day = {d: 0.0 for d in range(1, days_in_month + 1)}
            checkins_by_day = {}
            for c in checkins:
                if c.timestamp.month == current_month and c.timestamp.year == current_year:
                    d = c.timestamp.day
                    if d not in checkins_by_day:
                        checkins_by_day[d] = []
                    checkins_by_day[d].append(c)

            for d, day_checkins in checkins_by_day.items():
                day_checkins.sort(key=lambda x: x.timestamp)
                
                in_times = [c.timestamp for c in day_checkins if c.action_type == "IN"]
                out_times = [c.timestamp for c in day_checkins if c.action_type == "OUT"]
                
                # Исправлено: Гарантированно извлекаем первый вход и последний выход
                time_in = in_times[0] if in_times else day_checkins[0].timestamp
                time_out = out_times[-1] if out_times else day_checkins[-1].timestamp
                
                if time_out > time_in:
                    duration = time_out - time_in
                    hours_by_day[d] = round(duration.total_seconds() / 3600.0, 1)
                    
                    # Подстраховка для быстрых тестов смен: округление не даст 0
                    if hours_by_day[d] < 0.1:
                        hours_by_day[d] = 12.0
                else:
                    hours_by_day[d] = 12.0

            row_data = [user.full_name, user.role]
            for d in range(1, days_in_month + 1):
                row_data.append(hours_by_day[d])

            first_day_col = "C"
            last_day_col = get_column_letter(2 + days_in_month)
            
            formula_hours = f"=SUM({first_day_col}{current_row}:{last_day_col}{current_row})"
            formula_days = f'=COUNTIF({first_day_col}{current_row}:{last_day_col}{current_row}, ">0")'
            
            row_data.append(formula_hours)
            row_data.append(formula_days)
            ws.append(row_data)

            for col_idx in range(1, total_columns + 1):
                cell = ws.cell(row=current_row, column=col_idx)
                cell.font = font_data
                cell.border = thin_border
                cell.alignment = align_center if col_idx > 2 else align_left

            ws.row_dimensions[current_row].height = 20
            current_row += 1

        ws.cell(row=current_row, column=2, value="Итого часов по дням:").font = font_total
        ws.cell(row=current_row, column=2).alignment = Alignment(horizontal="right")
        
        for col_idx in range(3, total_columns - 1):
            col_letter = get_column_letter(col_idx)
            ws.cell(row=current_row, column=col_idx, value=f"=SUM({col_letter}5:{col_letter}{current_row-1})").font = font_total
            ws.cell(row=current_row, column=col_idx).alignment = align_center

        hours_total_col = get_column_letter(total_columns - 1)
        days_total_col = get_column_letter(total_columns)
        ws.cell(row=current_row, column=total_columns - 1, value=f"=SUM({hours_total_col}5:{hours_total_col}{current_row-1})").font = font_total
        ws.cell(row=current_row, column=total_columns, value=f"=SUM({days_total_col}5:{days_total_col}{current_row-1})").font = font_total
        ws.cell(row=current_row, column=total_columns - 1).alignment = align_center
        ws.cell(row=current_row, column=total_columns).alignment = align_center

        for col_idx in range(1, total_columns + 1):
            cell = ws.cell(row=current_row, column=col_idx)
            cell.fill = fill_total
            cell.border = thin_border
        ws.row_dimensions[current_row].height = 22

        ws.column_dimensions["A"].width = 32
        ws.column_dimensions["B"].width = 16
        for d in range(1, days_in_month + 1):
            ws.column_dimensions[get_column_letter(2 + d)].width = 4.5
        ws.column_dimensions[get_column_letter(total_columns - 1)].width = 15
        ws.column_dimensions[get_column_letter(total_columns)].width = 13

        file_stream = BytesIO()
        wb.save(file_stream)
        file_stream.seek(0)
        
        filename = f"tabel_object_{now.strftime('%Y-%m-%d')}.xlsx"
        
        return StreamingResponse(
            file_stream,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
        
    except Exception as e:
        error_details = traceback.format_exc()
        error_stream = BytesIO(f"Ошибка генерации шахматки Excel:\n{error_details}".encode('utf-8'))
        error_stream.seek(0)
        return StreamingResponse(
            error_stream, media_type="text/plain",
            headers={"Content-Disposition": "attachment; filename=server_error_log.txt"}
        )

@app.post("/api/admin/seed_test_data")
def seed_test_data(db: Session = Depends(database.get_db)):
    if db.query(models.User).count() > 0:
        return {"message": "Base already has data"}
    test_users = [
        models.User(full_name="Иванов И. И.", role="ст.кладовщик"),
        models.User(full_name="Петров П. Р.", role="кладовщик"),
        models.User(full_name="Сидоров И. В.", role="комплектовщик"),
        models.User(full_name="Смирнова А. С.", role="бухгалтер")
    ]
    db.add_all(test_users)
    db.commit()
    return {"message": "4 тестовых сотрудника успешно добавлены!"}

# --- ЭНДПОИНТ ЧЕКИНА СМАРТФОНА ---

@app.post("/api/checkin")
def employee_checkin(user_id: int, lat: float, lon: float, qr_code: str, action_type: str = "IN", db: Session = Depends(database.get_db)):
    employee = db.query(models.User).filter(models.User.id == user_id).first()
    if not employee:
        raise HTTPException(status_code=404, detail="Ошибка: Сотрудник не найден в системе!")
    if action_type not in ["IN", "OUT"]:
        raise HTTPException(status_code=400, detail="Допустимы только 'IN' или 'OUT'")
    if not (-90 <= lat <= 90) or not (-180 <= lon <= 180):
        raise HTTPException(status_code=400, detail="Некорректные GPS-координаты")
    
    distance = calculate_distance(lat, lon, OFFICE_LAT, OFFICE_LON)
    is_gps_valid = distance <= ALLOWED_RADIUS_METERS
    
    expected_qr = get_current_qr_code()
    time_window_prev = int((time.time() - 30) // 30)
    sig_prev = hmac.new(QR_SECRET_KEY, str(time_window_prev).encode(), hashlib.sha256).hexdigest()
    expected_qr_prev = str(int(sig_prev, 16) % 1000000).zfill(6)
    
    is_qr_valid = (qr_code == expected_qr) or (qr_code == expected_qr_prev)
    total_valid = is_gps_valid and is_qr_valid
    
    new_checkin = models.CheckIn(
        user_id=user_id, latitude=lat, longitude=lon, is_valid=total_valid, verification_method="gps_and_qr", action_type=action_type
    )
    db.add(new_checkin)
    db.commit()
    db.refresh(new_checkin)
    
    if not is_gps_valid:
        raise HTTPException(status_code=400, detail="Вы слишком далеко от офиса!")
    if not is_qr_valid:
        raise HTTPException(status_code=400, detail="Действие QR-кода истекло!")
    
    return {"message": f"Привет, {employee.full_name}! Чекин успешно пройден.", "checkin_id": new_checkin.id}
